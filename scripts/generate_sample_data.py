#!/usr/bin/env python3
"""
Generate synthetic sample data for Leadership Autopilot demo.

Creates fake but realistic-looking business metric Excel files
in the expected directory structure. All data is entirely fictional.

Usage:
    pip3 install openpyxl
    python3 scripts/generate_sample_data.py
"""

import os
import random
from openpyxl import Workbook

# Output directory
BASE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "sample")

# Fake product categories and subcategories
CATEGORIES = {
    "Smart Home": {
        "prefix": "1010",
        "subcats": [
            ("1001", "Voice Assistants"),
            ("1002", "Smart Lighting"),
            ("1003", "Smart Plugs"),
            ("1004", "Smart Thermostats"),
            ("1005", "Security Cameras"),
            ("1006", "Smart Doorbells"),
            ("1007", "Smart Speakers"),
            ("1008", "Home Automation Hubs"),
        ],
    },
    "Fitness Gear": {
        "prefix": "1020",
        "subcats": [
            ("2001", "Resistance Bands"),
            ("2002", "Yoga Mats"),
            ("2003", "Dumbbells"),
            ("2004", "Exercise Bikes"),
            ("2005", "Jump Ropes"),
            ("2006", "Foam Rollers"),
        ],
    },
    "Kitchen Gadgets": {
        "prefix": "1030",
        "subcats": [
            ("3001", "Air Fryers"),
            ("3002", "Blenders"),
            ("3003", "Coffee Makers"),
            ("3004", "Instant Pots"),
            ("3005", "Food Processors"),
            ("3006", "Toasters"),
            ("3007", "Electric Kettles"),
            ("3008", "Sous Vide Cookers"),
            ("3009", "Waffle Makers"),
            ("3010", "Slow Cookers"),
        ],
    },
    "Pet Tech": {
        "prefix": "1040",
        "subcats": [
            ("4001", "Automatic Feeders"),
            ("4002", "GPS Trackers"),
            ("4003", "Smart Pet Doors"),
            ("4004", "Pet Cameras"),
            ("4005", "Self-Cleaning Litter Boxes"),
        ],
    },
    "Gaming Accessories": {
        "prefix": "1050",
        "subcats": [
            ("5001", "Gaming Headsets"),
            ("5002", "Gaming Mice"),
            ("5003", "Mechanical Keyboards"),
            ("5004", "Controller Grips"),
            ("5005", "RGB LED Strips"),
            ("5006", "Streaming Microphones"),
            ("5007", "Monitor Stands"),
        ],
    },
}

# Fake ASINs per subcategory (3-5 per subcat)
def generate_fake_asins(subcat_code, count=4):
    return [f"B0FAKE{subcat_code}{i:02d}" for i in range(1, count + 1)]


# Metrics to generate
STANDARD_METRICS = ["GMS", "ShippedUnits"]
MARGIN_METRICS = ["ASP", "NetPPMLessSD", "CM", "SOROOS_PROCURABLE_PRODUCT_OOS_GV_PCT"]
ALL_METRICS = STANDARD_METRICS + MARGIN_METRICS

# Seed for reproducibility
random.seed(42)


def rand_value(base, spread=0.3):
    """Generate a random value around a base with some spread."""
    return base * (1 + random.uniform(-spread, spread))


def generate_standard_data(subcats, metric, week_num):
    """Generate data for standard layout metrics (GMS, ShippedUnits)."""
    rows = []
    total_value = 0
    total_wow_ctc = 0
    total_yoy_ctc = 0

    for code, name in subcats:
        if metric == "GMS":
            value = rand_value(random.uniform(50000, 500000))
        else:  # ShippedUnits
            value = rand_value(random.uniform(1000, 50000))

        wow_pct = random.uniform(-0.15, 0.20)
        yoy_pct = random.uniform(-0.30, 0.80)

        prior_wow = value / (1 + wow_pct)
        prior_yoy = value / (1 + yoy_pct)

        wow_ctc_dollars = value - prior_wow
        yoy_ctc_dollars = value - prior_yoy

        total_value += value
        total_wow_ctc += wow_ctc_dollars
        total_yoy_ctc += yoy_ctc_dollars

        rows.append({
            "code": code,
            "name": name,
            "value": round(value, 2),
            "wow_pct": round(wow_pct, 4),
            "yoy_pct": round(yoy_pct, 4),
            "wow_ctc_dollars": round(wow_ctc_dollars, 2),
            "wow_ctc_bps": 0,  # computed after totals
            "yoy_ctc_dollars": round(yoy_ctc_dollars, 2),
            "yoy_ctc_bps": 0,
        })

    # Compute bps CTC relative to totals
    total_prior_wow = total_value - total_wow_ctc
    total_prior_yoy = total_value - total_yoy_ctc
    total_wow_pct = total_wow_ctc / total_prior_wow if total_prior_wow else 0
    total_yoy_pct = total_yoy_ctc / total_prior_yoy if total_prior_yoy else 0

    for row in rows:
        if total_wow_ctc != 0:
            row["wow_ctc_bps"] = round(
                (row["wow_ctc_dollars"] / total_wow_ctc) * total_wow_pct * 10000
            )
        if total_yoy_ctc != 0:
            row["yoy_ctc_bps"] = round(
                (row["yoy_ctc_dollars"] / total_yoy_ctc) * total_yoy_pct * 10000
            )

    total_row = {
        "code": "Total",
        "name": "",
        "value": round(total_value, 2),
        "wow_pct": round(total_wow_pct, 4),
        "yoy_pct": round(total_yoy_pct, 4),
        "wow_ctc_dollars": round(total_wow_ctc, 2),
        "wow_ctc_bps": round(total_wow_pct * 10000),
        "yoy_ctc_dollars": round(total_yoy_ctc, 2),
        "yoy_ctc_bps": round(total_yoy_pct * 10000),
    }

    return rows, total_row


def generate_margin_data(subcats, metric):
    """Generate data for margin layout metrics (ASP, NetPPMLessSD, CM, SOROOS)."""
    rows = []

    for code, name in subcats:
        if metric == "ASP":
            value = rand_value(random.uniform(15, 200))
            numerator = value * rand_value(random.uniform(500, 5000))
            denominator = numerator / value
        elif metric == "NetPPMLessSD":
            value = random.uniform(0.05, 0.45)
            denominator = rand_value(random.uniform(50000, 500000))
            numerator = value * denominator
        elif metric == "CM":
            value = random.uniform(-0.05, 0.15)
            denominator = rand_value(random.uniform(50000, 500000))
            numerator = value * denominator
        else:  # SOROOS
            value = random.uniform(0.01, 0.15)
            denominator = rand_value(random.uniform(50000, 500000))
            numerator = value * denominator

        if metric == "ASP":
            wow_change = random.uniform(-0.10, 0.10)
            yoy_change = random.uniform(-0.20, 0.30)
        else:
            wow_change = random.uniform(-200, 200)  # bps
            yoy_change = random.uniform(-500, 500)  # bps

        wow_ctc = random.uniform(-50, 50) if metric != "ASP" else random.uniform(-5, 5)
        wow_mix = wow_ctc * random.uniform(0.3, 0.7)
        wow_rate = wow_ctc - wow_mix

        yoy_ctc = random.uniform(-150, 150) if metric != "ASP" else random.uniform(-15, 15)
        yoy_mix = yoy_ctc * random.uniform(0.3, 0.7)
        yoy_rate = yoy_ctc - yoy_mix

        rows.append({
            "code": code,
            "name": name,
            "value": round(value, 4) if metric != "ASP" else round(value, 2),
            "numerator": round(numerator, 2),
            "denominator": round(denominator, 2),
            "wow_change": round(wow_change, 4) if metric == "ASP" else round(wow_change),
            "yoy_change": round(yoy_change, 4) if metric == "ASP" else round(yoy_change),
            "wow_ctc": round(wow_ctc, 2),
            "wow_mix": round(wow_mix, 2),
            "wow_rate": round(wow_rate, 2),
            "yoy_ctc": round(yoy_ctc, 2),
            "yoy_mix": round(yoy_mix, 2),
            "yoy_rate": round(yoy_rate, 2),
        })

    # Compute totals
    total_num = sum(r["numerator"] for r in rows)
    total_den = sum(r["denominator"] for r in rows)
    total_value = total_num / total_den if total_den else 0

    total_row = {
        "code": "Total",
        "name": "",
        "value": round(total_value, 4) if metric != "ASP" else round(total_value, 2),
        "numerator": round(total_num, 2),
        "denominator": round(total_den, 2),
        "wow_change": round(sum(r["wow_ctc"] for r in rows), 2),
        "yoy_change": round(sum(r["yoy_ctc"] for r in rows), 2),
        "wow_ctc": round(sum(r["wow_ctc"] for r in rows), 2),
        "wow_mix": round(sum(r["wow_mix"] for r in rows), 2),
        "wow_rate": round(sum(r["wow_rate"] for r in rows), 2),
        "yoy_ctc": round(sum(r["yoy_ctc"] for r in rows), 2),
        "yoy_mix": round(sum(r["yoy_mix"] for r in rows), 2),
        "yoy_rate": round(sum(r["yoy_rate"] for r in rows), 2),
    }

    return rows, total_row


def generate_asin_data(subcats, metric, is_standard):
    """Generate ASIN-level data for a metric."""
    rows = []

    for code, name in subcats:
        asins = generate_fake_asins(code, random.randint(3, 5))
        for asin in asins:
            asin_name = f"Sample Brand {name} - Model {asin[-2:]}"

            if is_standard:
                if metric == "GMS":
                    value = rand_value(random.uniform(5000, 100000))
                else:
                    value = rand_value(random.uniform(100, 10000))

                wow_pct = random.uniform(-0.15, 0.20)
                yoy_pct = random.uniform(-0.30, 0.80)
                wow_ctc_d = value * wow_pct / (1 + wow_pct)
                yoy_ctc_d = value * yoy_pct / (1 + yoy_pct)

                rows.append({
                    "code": asin,
                    "name": asin_name,
                    "value": round(value, 2),
                    "wow_pct": round(wow_pct, 4),
                    "yoy_pct": round(yoy_pct, 4),
                    "wow_ctc_dollars": round(wow_ctc_d, 2),
                    "wow_ctc_bps": round(random.uniform(-100, 200)),
                    "yoy_ctc_dollars": round(yoy_ctc_d, 2),
                    "yoy_ctc_bps": round(random.uniform(-200, 500)),
                })
            else:
                if metric == "ASP":
                    value = rand_value(random.uniform(15, 200))
                elif metric == "NetPPMLessSD":
                    value = random.uniform(0.05, 0.45)
                elif metric == "CM":
                    value = random.uniform(-0.05, 0.15)
                else:
                    value = random.uniform(0.01, 0.15)

                numerator = value * rand_value(random.uniform(5000, 50000))
                denominator = numerator / value if value else 1

                rows.append({
                    "code": asin,
                    "name": asin_name,
                    "value": round(value, 4) if metric != "ASP" else round(value, 2),
                    "numerator": round(numerator, 2),
                    "denominator": round(denominator, 2),
                    "wow_change": round(random.uniform(-0.10, 0.10), 4) if metric == "ASP" else round(random.uniform(-200, 200)),
                    "yoy_change": round(random.uniform(-0.20, 0.30), 4) if metric == "ASP" else round(random.uniform(-500, 500)),
                    "wow_ctc": round(random.uniform(-50, 50), 2),
                    "wow_mix": round(random.uniform(-25, 25), 2),
                    "wow_rate": round(random.uniform(-25, 25), 2),
                    "yoy_ctc": round(random.uniform(-150, 150), 2),
                    "yoy_mix": round(random.uniform(-75, 75), 2),
                    "yoy_rate": round(random.uniform(-75, 75), 2),
                })

    return rows


def write_standard_xlsx(filepath, rows, total_row):
    """Write a standard layout (9-column) Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header
    ws.append(["Code", "Description", "Value", "WoW %", "YoY %",
               "WoW CTC ($)", "WoW CTC (bps)", "YoY CTC ($)", "YoY CTC (bps)"])

    # Total row first
    ws.append([
        total_row["code"], total_row["name"], total_row["value"],
        total_row["wow_pct"], total_row["yoy_pct"],
        total_row["wow_ctc_dollars"], total_row["wow_ctc_bps"],
        total_row["yoy_ctc_dollars"], total_row["yoy_ctc_bps"],
    ])

    # Data rows
    for row in rows:
        ws.append([
            row["code"], row["name"], row["value"],
            row["wow_pct"], row["yoy_pct"],
            row["wow_ctc_dollars"], row["wow_ctc_bps"],
            row["yoy_ctc_dollars"], row["yoy_ctc_bps"],
        ])

    wb.save(filepath)


def write_margin_xlsx(filepath, rows, total_row):
    """Write a margin layout (13-column) Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header
    ws.append(["Code", "Description", "Value", "Numerator ($)", "Denominator ($)",
               "WoW", "YoY", "WoW CTC", "WoW Mix", "WoW Rate",
               "YoY CTC", "YoY Mix", "YoY Rate"])

    # Total row first
    ws.append([
        total_row["code"], total_row["name"], total_row["value"],
        total_row["numerator"], total_row["denominator"],
        total_row["wow_change"], total_row["yoy_change"],
        total_row["wow_ctc"], total_row["wow_mix"], total_row["wow_rate"],
        total_row["yoy_ctc"], total_row["yoy_mix"], total_row["yoy_rate"],
    ])

    # Data rows
    for row in rows:
        ws.append([
            row["code"], row["name"], row["value"],
            row["numerator"], row["denominator"],
            row["wow_change"], row["yoy_change"],
            row["wow_ctc"], row["wow_mix"], row["wow_rate"],
            row["yoy_ctc"], row["yoy_mix"], row["yoy_rate"],
        ])

    wb.save(filepath)


def write_asin_standard_xlsx(filepath, rows):
    """Write ASIN-level standard layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ASIN", "Item Name", "Value", "WoW %", "YoY %",
               "WoW CTC ($)", "WoW CTC (bps)", "YoY CTC ($)", "YoY CTC (bps)"])

    for row in rows:
        ws.append([
            row["code"], row["name"], row["value"],
            row["wow_pct"], row["yoy_pct"],
            row["wow_ctc_dollars"], row["wow_ctc_bps"],
            row["yoy_ctc_dollars"], row["yoy_ctc_bps"],
        ])

    wb.save(filepath)


def write_asin_margin_xlsx(filepath, rows):
    """Write ASIN-level margin layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ASIN", "Item Name", "Value", "Numerator ($)", "Denominator ($)",
               "WoW", "YoY", "WoW CTC", "WoW Mix", "WoW Rate",
               "YoY CTC", "YoY Mix", "YoY Rate"])

    for row in rows:
        ws.append([
            row["code"], row["name"], row["value"],
            row["numerator"], row["denominator"],
            row["wow_change"], row["yoy_change"],
            row["wow_ctc"], row["wow_mix"], row["wow_rate"],
            row["yoy_ctc"], row["yoy_mix"], row["yoy_rate"],
        ])

    wb.save(filepath)


def generate_manifest(week, week_num, metrics, out_dir):
    """Generate _manifest.yaml for the week."""
    lines = [
        f'gl: all',
        f'week: "{week}"',
        f'generated: "2099-01-{week_num:02d}T00:00:00"',
        '',
        'metrics_available:',
    ]
    for m in metrics:
        lines.append(f'  - {m}')
    lines.append('')
    lines.append('files:')
    lines.append('  subcat:')
    for m in metrics:
        lines.append(f'    {m}: {m}_Week {week_num}_ctc_by_SUBCAT.xlsx')
    lines.append('  asin:')
    for m in metrics:
        lines.append(f'    {m}: {m}_Week {week_num}_ctc_by_ASIN.xlsx')

    with open(os.path.join(out_dir, "_manifest.yaml"), "w") as f:
        f.write("\n".join(lines) + "\n")


def main():
    weeks = [
        ("2099-wk01", 1),
        ("2099-wk02", 2),
    ]

    # Collect all subcats across categories
    all_subcats = []
    for cat_name, cat_data in CATEGORIES.items():
        prefix = cat_data["prefix"]
        for code_suffix, name in cat_data["subcats"]:
            full_code = f"{prefix}{code_suffix}"
            all_subcats.append((full_code, name))

    for week, week_num in weeks:
        out_dir = os.path.join(BASE_DIR, week, "ALL")
        os.makedirs(out_dir, exist_ok=True)

        print(f"Generating data for {week}...")

        for metric in ALL_METRICS:
            is_standard = metric in STANDARD_METRICS

            # Subcat-level files
            if is_standard:
                rows, total = generate_standard_data(all_subcats, metric, week_num)
                filepath = os.path.join(out_dir, f"{metric}_Week {week_num}_ctc_by_SUBCAT.xlsx")
                write_standard_xlsx(filepath, rows, total)
            else:
                rows, total = generate_margin_data(all_subcats, metric)
                filepath = os.path.join(out_dir, f"{metric}_Week {week_num}_ctc_by_SUBCAT.xlsx")
                write_margin_xlsx(filepath, rows, total)

            # ASIN-level files
            asin_rows = generate_asin_data(all_subcats, metric, is_standard)
            asin_filepath = os.path.join(out_dir, f"{metric}_Week {week_num}_ctc_by_ASIN.xlsx")
            if is_standard:
                write_asin_standard_xlsx(asin_filepath, asin_rows)
            else:
                write_asin_margin_xlsx(asin_filepath, asin_rows)

            print(f"  ✓ {metric} (SUBCAT + ASIN)")

        # Generate manifest
        generate_manifest(week, week_num, ALL_METRICS, out_dir)
        print(f"  ✓ _manifest.yaml")

    print(f"\nSample data generated in {BASE_DIR}/")
    print("To use: copy data/sample/2099-wk01/ to data/weekly/2099-wk01/")


if __name__ == "__main__":
    main()
